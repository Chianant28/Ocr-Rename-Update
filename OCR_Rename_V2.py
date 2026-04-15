import sys, os, time, ctypes
from tkinter import *
from tkinter import ttk   # ✅ import ครั้งเดียวพอ

# ---------- resource path ----------
def resource_path(p):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, p)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), p)

# ซ่อน console (ตอนเป็น exe)
ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)

# ---------- ROOT (มีแค่ครั้งเดียว) ----------
root = Tk()
root.attributes("-alpha", 0)
root.overrideredirect(True)
root.update_idletasks()
root.withdraw()   # ✅ ซ่อน root

# ---------------- Splash ----------------
splash = Toplevel(root)
splash.overrideredirect(True)
splash.configure(bg="black")
splash.attributes("-alpha", 0.95)

# จัดตำแหน่งกลางจอ
sw = splash.winfo_screenwidth()
sh = splash.winfo_screenheight()
w, h = 520, 240
x = (sw - w) // 2
y = (sh - h) // 2
splash.geometry(f"{w}x{h}+{x}+{y}")

# ---------- Canvas ----------
canvas = Canvas(splash, width=w, height=h, bg="black", highlightthickness=0)
canvas.pack(fill="both", expand=True)

# ---------- ฟังก์ชันวาดมุมโค้ง ----------
def round_rect(x1, y1, x2, y2, r=30, **kwargs):
    points = [
        x1+r, y1,
        x2-r, y1,
        x2, y1,
        x2, y1+r,
        x2, y2-r,
        x2, y2,
        x2-r, y2,
        x1+r, y2,
        x1, y2,
        x1, y2-r,
        x1, y1+r,
        x1, y1
    ]
    return canvas.create_polygon(points, smooth=True, **kwargs)

# ---------- Glow Effect ----------
for i in range(6):
    canvas.create_rectangle(
        10-i, 10-i, w-10+i, h-10+i,
        outline="#14E7EB",
        width=1
    )

# ---------- กล่องหลัก ----------
round_rect(10, 10, w-10, h-10, r=35,
           fill="#0A1F2B", outline="#14E7EB", width=2)

# ---------- ข้อความ ----------
load_text = StringVar(value="Starting OCR Rename...")
text_id = canvas.create_text(
    w//2, h//2 - 30,
    text=load_text.get(),
    fill="#14EB18",
    font=("Segoe UI", 20, "bold")
)

# ---------- Progress Bar ----------
style = ttk.Style()
style.theme_use('default')
style.configure(
    "cyan.Horizontal.TProgressbar",
    troughcolor="#0A1F2B",
    background="#14EB18",
    thickness=10
)

progress = ttk.Progressbar(
    splash,
    style="cyan.Horizontal.TProgressbar",
    orient="horizontal",
    length=300,
    mode="determinate"
)

canvas.create_window(w//2, h//2 + 40, window=progress)

# ---------- ฟังก์ชันอัปเดต ----------
def update_text(text, value):
    load_text.set(text)
    canvas.itemconfig(text_id, text=text)
    progress['value'] = value
    splash.update()

# ---------- Load ----------
update_text("Loading OpenCV...", 20)
import re, cv2, numpy as np

update_text("Loading OCR Engines...", 50)
import pytesseract, easyocr

update_text("Loading UI Libraries...", 75)
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook, Workbook
from rapidfuzz import process
import hashlib

update_text("Starting Application...", 100)

# ---------- Load icon ----------
#icon = PhotoImage(file=resource_path("image.png"))
#root.iconphoto(True, icon)

# ---------- ปิด Splash ----------
splash.after(500, splash.destroy)

# ---------- แสดงหน้าหลัก ----------
root.deiconify()
root.attributes("-alpha", 1)   # กลับมามองเห็น

#root.mainloop()
# ---------- Load icon ----------
icon = PhotoImage(file=resource_path("image.png"))
root.iconphoto(True, icon)

# ---------- Done ----------
#load_text.set("Ready")
#splash.update()
#time.sleep(0.5)
#splash.destroy()
#root.deiconify()

ai_cache = {}

def img_hash(img):
    return hashlib.md5(img.tobytes()).hexdigest()

def safe_gray(img):
    if img is None:
        return None
    if len(img.shape) == 2:          # already gray
        return img
    if img.shape[2] == 1:            # 1 channel
        return img[:,:,0]
    return cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

REF_FILE="ref_path.txt"
folder=""; files=[]
img_cv=None; img_tk=None
crop_box=None; start=None; temp_box=None
zoom=1.0; disp_w=disp_h=1
excel_map={}
offset_x=0; offset_y=0
pan_start=None
ok_count=0; fail_count=0
cancel_flag=False
pause_flag=False
view_locked=False
lock_zoom=1.0
lock_offset_x=0
lock_offset_y=0
img_base = None
redraw_pending = False
redraw_job = None
current_filename = ""
current_index = 0
NOTE_FILE = "Rename_ไม่สำเร็จ.txt"
note_list = []
raw_relation = None
rel_info = None
start_time = 0
dup_count = 0
running = False
total_files = 0
failed_files = []
dup_groups = {}
all_files = []
retry_mode = False
preview_mode = True
turbo_mode = False

BORDER = dict(highlightbackground="white", highlightthickness=2, bd=0)

def expand_crop(img, pad=0.18):
    h, w = img.shape[:2]
    pad_w = int(w * pad)
    pad_h = int(h * pad)
    # ถ้าเป็น Gray → แปลงเป็น BGR ก่อน
    if len(img.shape) == 2:
        img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    canvas = np.zeros((h + pad_h*2, w + pad_w*2, 3), dtype=np.uint8)
    canvas[pad_h:pad_h+h, pad_w:pad_w+w] = img
    return canvas

def set_rename_busy(busy):
    if busy:
        btn_rename.config(
            text="WORKING...",
            state="disabled",
            bg="#777777",
            fg="#dddddd"
        )
    else:
        btn_rename.config(
            text="RENAME",
            state="normal",
            bg="#00c853",
            fg="white"
        )

def export_folder_list_to_excel():
    global folder

    if not folder or not os.path.isdir(folder):
        messagebox.showerror("Error", "ยังไม่ได้เลือกโฟลเดอร์ภาพครับ")
        return

    files = []
    for f in os.listdir(folder):
        if f.lower().endswith((".jpg",".jpeg",".png",".bmp",".tif",".tiff",".webp")):
            files.append(f)

    if not files:
        messagebox.showinfo("Empty", "ไม่พบไฟล์ภาพในโฟลเดอร์นี้")
        return

    files.sort()

    save_path = filedialog.asksaveasfilename(
        title="Save Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel file","*.xlsx")],
        initialfile="รายชื่อระวาง.xlsx"
    )

    if not save_path:
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Renamed files"

    # ---------- Header ----------
    ws.append(["No", "Filename"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    # ---------- Data ----------
    for i, f in enumerate(files, start=1):
        name_only = os.path.splitext(f)[0]   # ตัด .jpg
        try:
            name_value = int(name_only)

        except:
            name_value = name_only
        ws.append([i, name_value])
        row = ws.max_row
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    total = len(files)

    # ---------- เว้น 1 บรรทัด ----------
    ws.append([])

    # ---------- Footer ----------
    row = ws.max_row + 2
    cell = ws.cell(row=row, column=2)
    cell.value = f"รวมจำนวน {total} ระวาง"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal= "center")

    # ---------- Auto width ----------
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(save_path)

    messagebox.showinfo("Done", f"บันทึก Excel แล้ว:\n{save_path}")

def show_big_message(title, text):
    win = Toplevel(root)
    win.title(title)
    win.configure(bg="#1e1e1e")
    win.grab_set()

    # ขยายขนาดให้พอดีกับข้อความ
    w, h = 800, 350
    sw = win.winfo_screenwidth()
    sh = win.winfo_screenheight()
    win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    msg = Label(
        win,
        text=text,
        fg="yellow",
        bg="#1e1e1e",
        font=("Segoe UI", 26, "bold"),
        wraplength=760,      # ให้ตัดบรรทัดอัตโนมัติ
        justify="center"
    )
    msg.pack(expand=True, padx=20, pady=30)

    Button(
        win,
        text="OK",
        font=("Segoe UI", 18, "bold"),
        bg="#00c853",
        fg="white",
        relief="flat",
        width=10,
        command=win.destroy
    ).pack(pady=15)

def normalize_retry_groups():
    global files
    tmp = []
    i = 1
    for base, flist in dup_groups.items():
        for f in flist:
            old = os.path.join(folder, f)
            if os.path.exists(old):
                new = os.path.join(folder, f"__retry_{i}.jpg")
                os.rename(old, new)
                tmp.append(f"__retry_{i}.jpg")
                i += 1
    files = tmp.copy()

def start_retry_mode():
    global files, retry_mode, ok_count, fail_count, dup_count, failed_files, total_files
    global all_files, current_index

    retry_mode = True

    # -------- 1. รวม FAIL + DUP --------
    retry_groups = []

    for f in failed_files:
        if os.path.exists(os.path.join(folder, f)):
            retry_groups.append(f)

    for base, flist in dup_groups.items():
        for f in flist:
            if os.path.exists(os.path.join(folder, f)):
                retry_groups.append(f)

    retry_groups = list(set(retry_groups))

    if not retry_groups:
        messagebox.showinfo("Retry", "ไม่มีไฟล์ที่ต้อง Retry")
        retry_mode = False
        return

    # -------- 2. NORMALIZE (ลบชื่อเก่า) --------
    files = []
    i = 1

    for f in retry_groups:
        old = os.path.join(folder, f)
        new = os.path.join(folder, f"__retry_{i}.jpg")

        try:
            os.rename(old, new)
            files.append(f"__retry_{i}.jpg")
            i += 1
        except:
            pass

    if not files:
        messagebox.showerror("Retry", "ไม่สามารถเตรียมไฟล์ Retry ได้")
        retry_mode = False
        return

    # ⭐ รีเซ็ต index และไฟล์ให้ UI ใช้ชุด retry
    current_index = 0
    all_files = files.copy()

    # -------- 3. Reset state --------
    total_files = len(files)
    ok_count = 0
    dup_count = 0
    fail_count = total_files

    failed_files.clear()
    dup_groups.clear()

    update_status()

    # -------- 4. บังคับ AI OCR --------
    ocr_mode.set("AI")   # หรือ ocr_mode = "AI"
    update_ocr_colors()
    btn_retry.pack_forget()

    show_big_message(
        "RETRY MODE",
        "ระบบรีเซ็ตชื่อไฟล์ซ้ำแล้ว\nกรุณาครอปภาพ แล้วกด RENAME"
    )

    # -------- 5. Preview ภาพแรก --------
    set_current_image(os.path.join(folder, all_files[0]))
    schedule_redraw()

    log("===== RETRY MODE (NORMALIZED) =====", "warn")
    log(f"ไฟล์ที่ต้องแก้ไข {total_files} ไฟล์", "warn")
    log("กรุณาครอปภาพแล้วกด RENAME อีกครั้ง", "warn")

def show_excel_path_only():
    if os.path.exists(REF_FILE):
        try:
            with open(REF_FILE, "r", encoding="utf8") as f:
                path = f.read().strip()
                if os.path.exists(path):
                    lbl_excel.config(text=os.path.basename(path))
                else:
                    lbl_excel.config(text="Excel not found")
        except:
            lbl_excel.config(text="No Excel")
    else:
        lbl_excel.config(text="No Excel")

def update_timer():
    if running:
        t = time.time() - start_time
        h = int(t // 3600)
        m = int((t % 3600) // 60)
        s = int(t % 60)
        timer_lbl.config(text=f"ใช้เวลา: {h:02d}:{m:02d}:{s:02d}")
    root.after(100, update_timer)

# ---------- LOG ----------
def log(msg, tag=None):
    # ให้ภาพถูกวาดเสร็จก่อน log
    root.update_idletasks()
    root.update()

    logbox.insert(END, msg+"\n", tag)
    logbox.see(END)

def sep(): log("-"*100)

# ---------- utils ----------
def cv_imread_th(path):
    data=np.fromfile(path,dtype=np.uint8)
    return cv2.imdecode(data,cv2.IMREAD_COLOR)

def ocr(img):
    g = safe_gray(img)
    g = cv2.resize(g,None,fx=2,fy=2)
    return pytesseract.image_to_string(g,lang="tha+eng").strip().upper()

def ocr_ai(crop):
    global ai_reader

    if ai_reader is None:
        return ""

    # ---------- Preprocess ----------
    gray = safe_gray(crop)
    gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)
    gray = cv2.GaussianBlur(gray, (3,3), 0)

    # ---------- Cache ----------
    h = img_hash(gray)
    if h in ai_cache:
        return ai_cache[h]

    # ---------- EasyOCR ----------
    result = ai_reader.readtext(gray, detail=0)

    text = " ".join(result).upper().strip()

    ai_cache[h] = text
    return text

def show_image(img):
    global img_tk, disp_w, disp_h
    rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    pil = Image.fromarray(rgb)

    disp_w, disp_h = pil.size
    img_tk = ImageTk.PhotoImage(pil)

    canvas.delete("img")
    canvas.delete("box")

    canvas.create_image(offset_x, offset_y, anchor="nw", image=img_tk, tags="img")

    if temp_box:
        canvas.create_rectangle(
            temp_box[0]+offset_x, temp_box[1]+offset_y,
            temp_box[2]+offset_x, temp_box[3]+offset_y,
            outline="red", width=2, tags="box"
        )

    if crop_box:
        canvas.create_rectangle(
            crop_box[0]+offset_x, crop_box[1]+offset_y,
            crop_box[2]+offset_x, crop_box[3]+offset_y,
            outline="yellow", width=3, tags="box"
        )

def redraw():
    global redraw_job
    if img_base is None:
        return

    redraw_job = None

    h, w = img_base.shape[:2]
    scale = zoom
    new = cv2.resize(img_base, (int(w*scale), int(h*scale)))

    show_image(new)

    # ----- TITLE FIX -----
    canvas.delete("title")
    canvas.delete("title_bg")

    cw = canvas.winfo_width()

    font_title = ("Segoe UI", 22, "bold")

    # วัดขนาดข้อความจริง
    tmp = canvas.create_text(0, 0, text=current_filename, font=font_title, anchor="nw")
    bbox = canvas.bbox(tmp)
    canvas.delete(tmp)

    if bbox:
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
    else:
        text_w = 200
        text_h = 30

    pad = 10
    x1 = cw//2 - text_w//2 - pad
    y1 = 4
    x2 = cw//2 + text_w//2 + pad
    y2 = y1 + text_h + pad

    # กล่องพื้นหลัง
    canvas.create_rectangle(
        x1, y1, x2, y2,
        fill="black",
        outline="",
        tags="title_bg"
    )

    # ตัวอักษร
    canvas.create_text(
        cw // 2,
        y1 + (text_h + pad)//2,
        text=current_filename,
        fill="cyan",
        font=font_title,
        anchor="center",
        tags="title"
    )

def schedule_redraw():
    global redraw_job
    if redraw_job:
        root.after_cancel(redraw_job)
    redraw_job = root.after(10, redraw)

# ---------- ZOOM ----------
def zoom_wheel(e):
    global zoom, offset_x, offset_y
    old_zoom = zoom
    zoom *= 1.1 if e.delta > 0 else 0.9
    zoom = max(0.1, min(zoom, 5))
    if img_cv is None: return

    cx = canvas.canvasx(e.x)
    cy = canvas.canvasy(e.y)
    img_x = (cx - offset_x) / old_zoom
    img_y = (cy - offset_y) / old_zoom
    offset_x = cx - img_x * zoom
    offset_y = cy - img_y * zoom
    schedule_redraw()

# ---------- PAN ----------
def pan_start_fn(e):
    global pan_start
    pan_start=(e.x,e.y)

def pan_move_fn(e):
    global offset_x, offset_y, pan_start
    if pan_start:
        dx = e.x - pan_start[0]
        dy = e.y - pan_start[1]

        offset_x += dx
        offset_y += dy

        pan_start = (e.x, e.y)
        schedule_redraw()

def pan_end_fn(e):
    global pan_start
    pan_start=None

def set_current_image(path):
    global img_cv, img_base, current_filename

    img = cv_imread_th(path)
    if img is None:
        return

    img_cv = img
    img_base = img.copy()
    current_filename = os.path.basename(path)

    schedule_redraw()


# ---------- FOLDER ----------
def choose_folder():
    global folder, files, current_index, zoom, offset_x, offset_y
    global ok_count, fail_count, dup_count, total_files, all_files
    global retry_mode

    folder = filedialog.askdirectory()

    # 🔥 RESET RETRY MODE
    retry_mode = False
    failed_files.clear()
    dup_groups.clear()
    all_files.clear()
    btn_retry.pack_forget()

    logbox.delete("1.0", END)

    ok_count = 0
    fail_count = 0
    dup_count = 0
    update_status()

    if not folder:
        return

    lbl_path.config(text=folder)
    log(f"เลือกโฟลเดอร์: {folder}")

    files = [f for f in os.listdir(folder) if f.lower().endswith(".jpg")]
    all_files = files.copy()
    total_files = len(files)

    ok_count = 0
    fail_count = 0
    dup_count = 0
    lbl_ok.config(text="0")
    lbl_fail.config(text="0")
    lbl_dup.config(text="0")

    files.sort()

    if not files:
        log("[ERROR] ไม่พบไฟล์ .jpg","fail_big")
        return

    current_index = 0

    # โหลดภาพแรก
    set_current_image(os.path.join(folder, files[0]))

    root.update_idletasks()
    cw = canvas.winfo_width()
    ch = canvas.winfo_height()

    h, w = img_base.shape[:2]

    # ซูมให้กว้างพอดี canvas
    zoom = cw / w

    # ขอบบนภาพ = ขอบบนจอ
    offset_x = 0
    offset_y = 0

    schedule_redraw()
    update_status()

# ---------- EXCEL ----------
def save_ref(path):
    with open(REF_FILE,"w",encoding="utf8") as f:f.write(path)

def load_excel_auto():
    if os.path.exists(REF_FILE):
        with open(REF_FILE,"r",encoding="utf8") as f:
            p=f.read().strip()
            if os.path.exists(p):
                load_excel(p)

def load_excel(path=None):
    global excel_map
    if not path:
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if not path:
            return
    lbl_excel.config(text=path)
    save_ref(path)
    log(f"โหลด Excel: {path}")
    wb = load_workbook(path)
    ws = wb.active
    excel_map = {}
    first = True
    for r in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue

        if r and r[0] and r[1]:
            excel_map[str(r[0]).strip()] = str(r[1]).strip()
    log(f"ระวางอ้างอิง {len(excel_map)} ระวาง")

# ---------- CROP ----------
def m_down(e):
    global start,temp_box
    start=(canvas.canvasx(e.x)-offset_x,canvas.canvasy(e.y)-offset_y)
    temp_box=None

def m_move(e):
    global temp_box
    if start:
        x1,y1=start
        x2,y2=(canvas.canvasx(e.x)-offset_x,
               canvas.canvasy(e.y)-offset_y)
        temp_box=(min(x1,x2),min(y1,y2),
                  max(x1,x2),max(y1,y2))
        schedule_redraw()   # 👈 ใช้ redraw แทน

def m_up(e):
    global crop_box,temp_box,start,view_locked,lock_zoom,lock_offset_x,lock_offset_y
    if not start:return
    crop_box=temp_box
    temp_box=None; start=None
    view_locked=True
    lock_zoom=zoom
    lock_offset_x=offset_x
    lock_offset_y=offset_y
    log("กำหนดพื้นที่ครอปแล้ว (ล็อกซูม/ตำแหน่ง)")
    schedule_redraw()      # 👈 แทน show_image

def show_crop_from(img):
    if not crop_box: return None
    h,w,_ = img.shape
    sx = w / disp_w
    sy = h / disp_h
    x1,y1,x2,y2 = crop_box
    return img[int(y1*sy):int(y2*sy),
               int(x1*sx):int(x2*sx)]

def full_preview():
    global zoom,offset_x,offset_y
    w = max(canvas.winfo_width(),1)
    zoom = w / img_base.shape[1]
    offset_x = offset_y = 0
    schedule_redraw()

def next_image():
    global current_index
    if not all_files:
        return

    current_index += 1
    if current_index >= len(all_files):
        current_index = len(all_files) - 1

    path = os.path.join(folder, all_files[current_index])
    set_current_image(path)
    schedule_redraw()

def prev_image():
    global current_index
    if not all_files:
        return

    current_index -= 1
    if current_index < 0:
        current_index = 0

    path = os.path.join(folder, all_files[current_index])
    set_current_image(path)
    schedule_redraw()

def update_status():
    lbl_total.config(text=str(total_files))   # 👈 คงที่ = จำนวนไฟล์จริง
    lbl_ok.config(text=str(ok_count))
    lbl_fail.config(text=str(fail_count))
    lbl_dup.config(text=str(dup_count))

# ---------- MATCH ----------
def match_by_probability(key):
    first4 = key[:4]
    last4  = key[-4:]
    front = [k for k in excel_map if k.startswith(first4)]
    if front: return front[0], "FRONT"
    back = [k for k in excel_map if k.endswith(last4)]
    if back: return back[0], "BACK"
    return None, None

def ai_match(key):
    keys = list(excel_map.keys())
    best = process.extractOne(key, keys, score_cutoff=70)
    if best:
        return best[0], best[1]
    return None, None

def log_fail(fname, reason):
    global fail_count
    log(f"[FAIL] {fname} : {reason}", "fail_big")
    note_list.append(f"{fname} -> FAIL : {reason}")
    fail_count += 1
    update_status()

def enhance_variants(img):
    gray = safe_gray(img)
    res = []
    res.append(gray)
    res.append(cv2.equalizeHist(gray))
    res.append(cv2.GaussianBlur(gray, (3,3), 0))
    res.append(cv2.threshold(gray,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1])
    res.append(cv2.adaptiveThreshold(
        gray,255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,11,2))

    return res

def ocr_both(img):
    g = safe_gray(img)
    img = cv2.cvtColor(g, cv2.COLOR_GRAY2BGR)
    t1 = pytesseract.image_to_string(img, lang="tha+eng").upper()
    t2 = ""
    if ai_reader:
        t2 = ocr_ai(img)
    return t1 + "\n" + t2

def smart_recheck(crop, use_ai=True):
    variants = []

    gray = safe_gray(crop)
    variants.append(gray)

    thr1 = cv2.adaptiveThreshold(gray,255,
                cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY,11,2)
    variants.append(thr1)

    thr2 = cv2.threshold(gray,0,255,
                cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1]
    variants.append(thr2)

    best_text = ""
    best_score = 0

    for v in variants:
        t1 = pytesseract.image_to_string(v, lang="tha+eng").upper()
        d1 = len(re.findall(r'\d', t1))

        if use_ai and ai_reader:
            t2 = ocr_ai(v)
            d2 = len(re.findall(r'\d', t2))
            if d2 > d1:
                t1 = t2
                d1 = d2

        if d1 > best_score:
            best_score = d1
            best_text = t1
    return best_text

def recheck_focus_mapno(crop, excel_map, use_ai=True):
    big = expand_crop(crop, 0.15)
    text = smart_recheck(big, use_ai)
    key = find_key_from_text(text, excel_map)

    if key:
        return key, text, True
    return None, text, False

def find_key_from_text(text, excel_map):
    groups = re.findall(r'\d+', text)

    for i in range(len(groups)):
        g1 = groups[i]

        if len(g1) < 4:
            continue

        # prefix จากกลุ่มหน้า
        prefixes = [g1[a:a+4] for a in range(len(g1) - 3)]

        for p in prefixes:
            # หาทุก key ใน Excel ที่ขึ้นต้นด้วย prefix นี้
            possible = [k for k in excel_map if k.startswith(p)]
            if not possible:
                continue

            # มองทุกกลุ่มหลังจากนี้เป็น suffix candidate
            for j in range(i + 1, len(groups)):
                g2 = groups[j]
                if len(g2) < 4:
                    continue

                suffixes = [g2[b:b+4] for b in range(len(g2) - 3)]

                for s in suffixes:
                    for k in possible:
                        if k.endswith(s):
                            return k

    return None

# ---------- RENAME ----------
def rename_all():
    global ok_count, fail_count, cancel_flag, img_cv, zoom, offset_x, offset_y
    global start_time, running, dup_count, retry_mode
    global files, total_files, all_files
    # 🔒 กันกดซ้ำ
    if running:
        return

    running = True
    set_rename_busy(True)     # ปุ่ม = WORKING…
    start_time = time.time()

    try:
        # ---------- PREPARE ----------
        if not retry_mode:
            failed_files.clear()
        else:
            log("===== RETRY MODE =====", "warn")

        if not excel_map:
            load_excel_auto()

        if not crop_box:
            messagebox.showerror("Error", "ยังไม่ได้ครอปภาพ")
            return

        if not excel_map:
            messagebox.showerror("Error", "ยังไม่ได้โหลด Excel")
            return

        cancel_flag = False
        ok_count = 0
        fail_count = 0
        dup_count = 0

        used = set(os.listdir(folder))
        log("====== START RENAME ======")

        note_list.clear()
        if os.path.exists(NOTE_FILE):
            os.remove(NOTE_FILE)

        # ---------- MAIN LOOP ----------
        run_list = all_files if retry_mode else files
        total_files = len(run_list)
        update_status()
        for f in list(run_list):   # 🔥 clone list เพื่อให้ลบใน loop ได้
            sep()
            last_ocr_text = ""
            mode = ocr_mode.get()   # "HYBRID" หรือ "AI"

            # ===== SKIP FILE IF ALREADY RENAMED =====
            name_only = os.path.splitext(f)[0]
            if re.fullmatch(r"\d{9}", name_only):
                log(f"[SKIP] {f} มีการ Rename ก่อนหน้านี้แล้ว")
                continue

            # Pause / Cancel
            root.update()
            while pause_flag:
                root.update()

            if cancel_flag:
                log("====== STOPPED BY USER ======", "fail_big")
                break

            try:
                path = os.path.join(folder, f)
                img = cv_imread_th(path)

                if preview_mode:
                    set_current_image(path)
                    schedule_redraw()
                    root.update_idletasks()
                    root.update()

                crop = show_crop_from(img)
                if crop is None:
                    log_fail(f, "ยังไม่ได้กำหนดพื้นที่ครอป")
                    failed_files.append(f)
                    continue

                path_mode = ""
                confidence = 100
                text = ""
                key = None

                # ---------- OCR + AI MODE ----------
                if mode == "HYBRID":
                    text = ocr(crop)
                    last_ocr_text = text
                    log(f"[OCR] {f} -> {text}")

                    key = find_key_from_text(text, excel_map)
                    path_mode = "OCR"
                    confidence = 100

                    if fast_recheck.get():
                        log("[RECHECK] Crop + Expand + Enhance...", "warn")
                        key2, t2, ok = recheck_focus_mapno(crop, excel_map)
                        last_ocr_text = t2
                        log(f"[RECHECK] {t2}")

                        if ok:
                            key = key2
                            path_mode = "OCR→RECHECK"
                            confidence = 98

                    if not key:
                        ai_text = ocr_ai(crop)
                        last_ocr_text = ai_text
                        log(f"[AI OCR] {ai_text}")
                        key = find_key_from_text(ai_text, excel_map)
                        path_mode = "OCR+AI"
                        confidence = 95

                # ---------- AI OCR MODE ----------
                elif mode == "AI":
                    log("[AI OCR MODE] ใช้ AI OCR", "warn")
                    ai_text = ocr_ai(crop)
                    last_ocr_text = ai_text
                    log(f"[AI OCR] {ai_text}")
                    key = find_key_from_text(ai_text, excel_map)
                    path_mode = "AI"
                    confidence = 100

                    if fast_recheck.get():
                        log("[AI RECHECK] Crop + Expand + Enhance...", "warn")
                        key2, t2, ok = recheck_focus_mapno(crop, excel_map)
                        last_ocr_text = t2
                        log(f"[AI RECHECK] {t2}")

                        if ok:
                            key = key2
                            path_mode = "AI→RECHECK"
                            confidence = 99

                # ---------- STILL FAIL ----------
                if not key:
                    msg = last_ocr_text.strip()
                    if not msg:
                        msg = ""
                    msg = msg + " (ไม่พบเลขระวางใด ๆ ที่ตรงกับระวางอ้างอิง)"
                    log_fail(f, msg)
                    failed_files.append(f)
                    continue

                base = excel_map[key]

                # ---------- RENAME ----------
                new = base + ".jpg"
                if new in used:
                    i = 1
                    original = new
                    while new in used:
                        new = f"{base}_ซ้ำ{i}.jpg"
                        i += 1

                    dup_count += 1

                    # 🔥 เก็บชื่อไฟล์จริงหลัง rename
                    if base not in dup_groups:
                        dup_groups[base] = []

                    # original = ชื่อไฟล์แรกที่ใช้ base นี้ (เช่น 493810048.jpg)
                    if original not in dup_groups[base]:
                        dup_groups[base].append(original)

                    # new = ชื่อไฟล์ที่ถูก rename เป็น _ซ้ำ1
                    if new not in dup_groups[base]:
                        dup_groups[base].append(new)

                    msg = f"{f} : DUPLICATE NAME -> {original} -> {new}"
                    note_list.append(msg)
                    logbox.insert(END, msg + "\n", "warn")
                    logbox.see(END)

                os.rename(path, os.path.join(folder, new))
                used.add(new)

                log(f"[OK] ({path_mode} {confidence}%) {f} -> {new}")
                ok_count += 1
                update_status()

                # 🔥 ถ้าเป็น retry → เอาไฟล์นี้ออกจาก all_files
                if retry_mode:
                    if f in all_files:
                        all_files.remove(f)

            except Exception as e:
                msg = last_ocr_text.strip()
                if not msg:
                    msg = ""
                msg = msg + " (ไม่พบเลขระวางใด ๆ ที่ตรงกับระวางอ้างอิง)"
                log(f"[FAIL] {f} : {msg}", "fail_big")
                note_list.append(f"{f} -> FAIL : {msg}")
                failed_files.append(f)
                fail_count += 1
                update_status()

        # ---------- FINISH ----------
        need_retry = False
        if retry_mode and all_files:
            need_retry = True
        elif failed_files:
            need_retry = True
        elif dup_groups:
            need_retry = True
        # ปุ่ม Retry
        if need_retry:
            btn_retry.pack(fill=X, pady=6)
        else:
            btn_retry.pack_forget()
        # Log สถานะ
        if need_retry:
            total_dup = sum(len(v) for v in dup_groups.values())
            total_left = len(all_files) if retry_mode else len(failed_files)
            log(
                f"เหลือไฟล์ต้องแก้ {total_left} | ไฟล์ซ้ำ {total_dup}",
                "warn"
            )
        else:
            log("====== FINISHED ======")

        # 🔔 popup ต้องขึ้นทุกครั้งที่ loop นี้จบ
        messagebox.showinfo("เสร็จ", f"สำเร็จ {ok_count} | ไม่สำเร็จ {fail_count}")

        # ออกจาก retry mode เมื่อไม่มีอะไรเหลือ
        if not need_retry:
            retry_mode = False

        # 🔥 ถ้ายังมี fail หรือ duplicate → ยังอยู่ใน retry mode
        if failed_files or dup_groups:
            retry_mode = True
        else:
            retry_mode = False

        # ---------- RETRY BUTTON LOGIC ----------
        need_retry = False

        if retry_mode and all_files:
            need_retry = True
        elif failed_files:
            need_retry = True
        elif dup_groups:
            need_retry = True

        if need_retry:
            btn_retry.pack(fill=X, pady=6)

            total_dup = sum(len(v) for v in dup_groups.values())
            total_left = len(all_files) if retry_mode else len(failed_files)

            log(
                f"เหลือไฟล์ต้องแก้ {total_left} | ไฟล์ซ้ำ {total_dup}  กด RENAME เพื่อแก้ไข",
                "warn"
            )
        else:
            btn_retry.pack_forget()

        if note_list:
            with open(NOTE_FILE, "w", encoding="utf8") as f:
                f.write("\n".join(note_list))
            try:
                os.startfile(NOTE_FILE)
            except:
                pass

        #messagebox.showinfo("เสร็จ", f"สำเร็จ {ok_count} | ไม่สำเร็จ {fail_count}")

    finally:
        # 🔓 ปลดปุ่ม + ปลด running (ต้องผ่าน UI thread)
        root.after(0, lambda: (
            set_rename_busy(False),
            globals().__setitem__("running", False)
        ))

# ---------- CANCEL ----------
def cancel_run():
    global cancel_flag, running
    cancel_flag=True
    running = False
    log("!!! USER CANCELLED !!!","fail_big")

def pause_run():
    global pause_flag
    pause_flag = not pause_flag

    if pause_flag:
        btn_pause.config(text="RESUME", bg="#00e5ff")
        log("|| PAUSED ||","warn")
    else:
        btn_pause.config(text="PAUSE", bg="#ffb300")
        log(">> RESUME <<","warn")

# ---------- LOAD OCR SYSTEM ----------
ai_reader = None
try:
    load_text.set("Loading EasyOCR model...")
    splash.update()

    ai_reader = easyocr.Reader(['th','en'], gpu=False)

    load_text.set("Checking Tesseract OCR...")
    splash.update()

except Exception as e:
    messagebox.showwarning(
        "EasyOCR Error",
        "ไม่สามารถโหลด EasyOCR ได้\n\n"
        "ตรวจสอบอินเทอร์เน็ต หรือ Firewall\n\n"
        + str(e)
    )
    ai_reader = None

load_text.set("Checking Tesseract OCR...")
splash.update()
TESS_PATH = resource_path("tesseract.exe")
if not os.path.exists(TESS_PATH):
    messagebox.showerror(
        "Missing Tesseract",
        "โปรแกรมนี้ต้องใช้ Tesseract OCR\n\n"
        "กรุณาติดตั้งก่อน:\n"
        "tesseract-ocr-w64-setup-5.5.0.exe\n\n"
        "แล้วเปิดโปรแกรมใหม่"
    )
    sys.exit()

pytesseract.pytesseract.tesseract_cmd = TESS_PATH

# ----- ปิด Splash และเปิดหน้าหลัก -----
load_text.set("Starting program...")
splash.update()
time.sleep(0.3)

splash.destroy()
root.overrideredirect(False)
root.attributes("-alpha", 1)     # กลับมามองเห็น
root.deiconify()
root.lift()
root.focus_force()

root.title("OCR Rename V2.0")
root.state("zoomed")
root.configure(bg="#1e1e1e")

def update_ocr_colors():
    if ocr_mode.get() == "HYBRID":
        rb_hybrid.config(fg="#00ff66")
        rb_ai.config(fg="white")
    else:
        rb_hybrid.config(fg="white")
        rb_ai.config(fg="#00ff66")

from PIL import Image, ImageTk

logo_img = Image.open(resource_path("image.png"))
logo_img = logo_img.resize((36,36), Image.LANCZOS)
logo_tk = ImageTk.PhotoImage(logo_img)

# ----- เริ่มสร้าง UI จริง -----
top = Frame(root, bg="#1e1e1e")
top.pack(fill=X, pady=1)
left_box  = Frame(top, bg="#1e1e1e")
mid_box   = Frame(top, bg="#1e1e1e")
right_box = Frame(top, bg="#1e1e1e")

left_box.pack(side=LEFT, padx=15, anchor="n")
# ---------- LOGO ----------
logo_frame = Frame(left_box, bg="#1e1e1e")
logo_frame.pack(anchor="w", pady=(4,8))
Label(logo_frame, image=logo_tk, bg="#1e1e1e").pack(side=LEFT)
Label(
    logo_frame,
    text="โปรแกรมสำหรับ Rename ภาพถ่ายแผนที่ด้วยวิธี Optical Character Recognition",
    fg="#ECF00F",
    bg="#1e1e1e",
    font=("Segoe UI",11,"bold"),
    justify="left"
).pack(side=LEFT, padx=8)
# -------------------------
mid_box.pack(side=LEFT, expand=True, anchor="n")
right_box.pack(side=RIGHT, padx=15, anchor="n")
# 👉 สร้างกรอบฝั่งขวาบนสำหรับ OCR MODE
# ---- OCR MODE PANEL (กรอบขาว) ----
ocr_panel = Frame(top, bg="#1e1e1e", highlightbackground="white", highlightthickness=2)
ocr_panel.pack(side=RIGHT, padx=15, pady=(0,5), anchor="n")

Label(ocr_panel, text="OCR MODE", fg="#00ffcc", bg="#1e1e1e",
      font=("Segoe UI",12,"bold")).pack(pady=(4,2))

ocr_mode = StringVar(value="HYBRID")      # 👈 Default = AI OCR
fast_recheck = BooleanVar(value=False)  # 👈 ปิด Recheck ตั้งแต่เริ่ม

rb_hybrid = Radiobutton(
    ocr_panel, text="OCR + AI (เร็ว)",
    variable=ocr_mode, value="HYBRID",
    bg="#1e1e1e", fg="#00ff66",   # ค่าเริ่มต้นเขียว
    selectcolor="#333",
    activebackground="#1e1e1e",
    font=("Segoe UI",12),
    command=update_ocr_colors
)
rb_hybrid.pack(anchor="w", padx=12)

rb_ai = Radiobutton(
    ocr_panel, text="AI OCR ONLY (แม่นกว่า)",
    variable=ocr_mode, value="AI",
    bg="#1e1e1e", fg="white",
    selectcolor="#333",
    activebackground="#1e1e1e",
    font=("Segoe UI",12),
    command=update_ocr_colors
)
rb_ai.pack(anchor="w", padx=12)

Checkbutton(
    ocr_panel,
    text="Recheck",
    variable=fast_recheck,
    bg="#1e1e1e",
    fg="cyan",
    selectcolor="#333",
    font=("Segoe UI", 17, "bold")
).pack(pady=(4,6))
update_ocr_colors()   # ⭐ sync สีเริ่มต้น
# --------- Folder + Reference Panel (2 rows) ---------
top_left = Frame(left_box, bg="#1e1e1e")
top_left.pack(anchor="w")

# Row 1 : Select Folder + Path
row_folder = Frame(top_left, bg="#1e1e1e")
row_folder.pack(anchor="w", pady=(0,6))

btn_select = Button(
    row_folder,
    text="Select Folder",
    font=("Segoe UI",18),
    bg="#007acc",
    fg="white",
    relief="raised",
    bd=4,
    width=12,
    command=choose_folder
)
btn_select.pack(side=LEFT)

lbl_path = Label(
    row_folder,
    fg="#00ffcc",
    bg="#1e1e1e",
    font=("Segoe UI",14,"bold"),
    anchor="w"
)
lbl_path.pack(side=LEFT, padx=(10,0))

# Row 2 : Reference + Excel path
row_ref = Frame(top_left, bg="#1e1e1e")
row_ref.pack(anchor="w")

btn_reference = Button(
    row_ref,
    text="Reference",
    width=13,
    bd=4,
    font=("Segoe UI",16),
    bg="#00c853",
    fg="black",
    relief="raised",
    command=load_excel
)
btn_reference.pack(side=LEFT)

lbl_excel = Label(
    row_ref,
    fg="#7CFC00",
    bg="#1e1e1e",
    font=("Segoe UI",14,"bold"),
    anchor="w"
)
lbl_excel.pack(side=LEFT, padx=(10,0))

main=Frame(root,bg="#1e1e1e"); main.pack(fill=BOTH,expand=True)

canvas=Canvas(main,bg="black",**BORDER)
canvas.pack(side=LEFT,fill=BOTH,expand=True,padx=10,pady=10)

right=Frame(main,width=420,bg="#2b2b2b"); right.pack(side=RIGHT,fill=BOTH)

top_right=Frame(right,bg="#2b2b2b",**BORDER); top_right.pack(fill=X,padx=10,pady=5)

log_frame=Frame(right,bg="#2b2b2b",**BORDER); log_frame.pack(fill=BOTH,expand=True,padx=10,pady=10)

canvas.bind("<Button-1>",m_down)
canvas.bind("<B1-Motion>",m_move)
canvas.bind("<ButtonRelease-1>",m_up)
canvas.bind("<MouseWheel>",zoom_wheel)
canvas.bind("<Button-3>",pan_start_fn)
canvas.bind("<B3-Motion>",pan_move_fn)
canvas.bind("<ButtonRelease-3>",pan_end_fn)

# FULL PREVIEW
btn_preview=Button(canvas,text="FULL PREVIEW",font=("Segoe UI",12,"bold"),
                   bg="#ff9800",fg="black",relief="flat",command=full_preview)
btn_preview.place(relx=1.0, rely=0.02, anchor="ne", x=-8, y=-10)

# NEXT
btn_next = Button(canvas, text="Next ▶", font=("Segoe UI",10,"bold"),
                  bg="#00c853", fg="white", relief="flat", command=next_image)
btn_next.place(relx=1.0, rely=0.02, anchor="ne", x=-150, y=-10)

# PREVIOUS
btn_prev = Button(canvas, text="◀ Previous", font=("Segoe UI",10,"bold"),
                  bg="#1976d2", fg="white", relief="flat", command=prev_image)
btn_prev.place(relx=1.0, rely=0.02, anchor="ne", x=-240, y=-10)

btn_rename = Button(
    top_right,
    text="RENAME",
    font=("Segoe UI",18,"bold"),
    bg="#00c853", fg="white",
    width=50,
    relief="raised",
    bd=4,
    highlightthickness=0,
    activebackground="#ff5252",
    activeforeground="white",
    command=rename_all
)
btn_rename.pack(padx=5, pady=5)

btn_retry = Button(
    top_right,
    text="RENAME ภาพที่ไม่สำเร็จอีกครั้ง",
    font=("Segoe UI",14,"bold"),
    bg="#ff1744",
    fg="white",
    relief="flat",
    command=lambda: start_retry_mode()
)

row_btn = Frame(top_right,bg="#2b2b2b")
row_btn.pack(fill=X,pady=5)

btn_pause = Button(
    row_btn,
    text="PAUSE",
    width=25,
    font=("Segoe UI",14,"bold"),
    bg="#ffb300",
    fg="black",
    relief="raised",     # ← ทำให้นูน
    bd=4,                # ← ความหนาขอบ
    highlightthickness=0,
    activebackground="#ff5252",   # สีตอนกด
    activeforeground="white",
    command=pause_run
)
btn_pause.pack(side=LEFT,expand=True,padx=2)

Button(
    row_btn,
    text="CANCEL",
    font=("Segoe UI",14,"bold"),
    bg="#e53935",
    fg="white",
    width=25,
    relief="raised",     # ← ทำให้นูน
    bd=4,                # ← ความหนาขอบ
    highlightthickness=0,

    activebackground="#ff5252",   # สีตอนกด
    activeforeground="white",

    command=cancel_run
).pack(side=LEFT, expand=True, padx=2)

status_frame = Frame(top_right, bg="#2b2b2b")
status_frame.pack(pady=3)

Button(status_frame,
       text="Export to Excel",
       command=export_folder_list_to_excel,
       bg="#4caf50",
       fg="white",
       font=("Segoe UI",10,"bold")
).pack(side=LEFT, padx=10, ipadx=10)

# ทั้งหมด
Label(status_frame, text="ทั้งหมด:", fg="#00e5ff", bg="#2b2b2b",
      font=("Segoe UI",12,"bold")).pack(side=LEFT, padx=(6,2))
lbl_total = Label(status_frame, text="0", fg="#00e5ff", bg="#2b2b2b",
                  font=("Segoe UI",12,"bold"))
lbl_total.pack(side=LEFT, padx=(0,10))

# สำเร็จ
Label(status_frame, text="สำเร็จ:", fg="#00ff6a", bg="#2b2b2b",
      font=("Segoe UI",12,"bold")).pack(side=LEFT, padx=(6,2))
lbl_ok = Label(status_frame, text="0", fg="#00ff6a", bg="#2b2b2b",
               font=("Segoe UI",12,"bold"))
lbl_ok.pack(side=LEFT, padx=(0,10))

# ไม่สำเร็จ
Label(status_frame, text="ไม่สำเร็จ:", fg="#00e5ff", bg="#2b2b2b",
      font=("Segoe UI",12,"bold")).pack(side=LEFT, padx=(6,2))
lbl_fail = Label(status_frame, text="0", fg="#ff1744", bg="#2b2b2b",
                 font=("Segoe UI",12,"bold"))
lbl_fail.pack(side=LEFT, padx=(0,10))

# ชื่อซ้ำ
Label(status_frame, text="ชื่อซ้ำ:", fg="#00e5ff", bg="#2b2b2b",
      font=("Segoe UI",12,"bold")).pack(side=LEFT, padx=(6,2))
lbl_dup = Label(status_frame, text="0", fg="#ff1744", bg="#2b2b2b",
                font=("Segoe UI",12,"bold"))
lbl_dup.pack(side=LEFT, padx=(0,10))

# ---- Mode Toggles ----
def update_mode_buttons():
    if preview_mode:
        btn_preview_mode.config(bg="#00c853", fg="white", relief="sunken")
        btn_turbo_mode.config(bg="#555", fg="white", relief="raised")
    else:
        btn_turbo_mode.config(bg="#00c853", fg="white", relief="sunken")
        btn_preview_mode.config(bg="#555", fg="white", relief="raised")

def set_preview():
    global preview_mode, turbo_mode
    preview_mode = True
    turbo_mode = False
    update_mode_buttons()

    if "logbox" in globals():
        log("MODE: PREVIEW","warn")

def set_turbo():
    global preview_mode, turbo_mode
    preview_mode = False
    turbo_mode = True
    update_mode_buttons()

    if "logbox" in globals():
        log("MODE: TURBO","warn")

btn_preview_mode = Button(
    status_frame,
    text="👁 PREVIEW",
    font=("Segoe UI",10,"bold"),
    command=set_preview
)
btn_preview_mode.pack(side=LEFT, padx=6)

btn_turbo_mode = Button(
    status_frame,
    text="🚀 TURBO",
    font=("Segoe UI",10,"bold"),
    command=set_turbo
)
btn_turbo_mode.pack(side=LEFT, padx=6)
set_preview()

timer_lbl = Label(
    top_right,
    text="ใช้เวลา: 00:00:00",
    fg="#00e5ff",
    bg="#2b2b2b",
    font=("Segoe UI",12,"bold")
)
timer_lbl.pack()
update_timer()

ys=Scrollbar(log_frame,orient=VERTICAL); ys.pack(side=RIGHT,fill=Y)
xs=Scrollbar(log_frame,orient=HORIZONTAL); xs.pack(side=BOTTOM,fill=X)

logbox=Text(log_frame,bg="black",fg="lime",font=("Consolas",12),
            yscrollcommand=ys.set,xscrollcommand=xs.set,wrap=NONE)
logbox.pack(fill=BOTH,expand=True)

ys.config(command=logbox.yview)
xs.config(command=logbox.xview)

logbox.tag_config("warn",foreground="yellow")
logbox.tag_config("fail_big",foreground="red",font=("Consolas",14,"bold"))
logbox.tag_config("relation", foreground="red", font=("Consolas",14,"bold"))
show_excel_path_only()

def on_close():
    if messagebox.askyesno("Exit", "ต้องการปิดโปรแกรมหรือไม่?"):
        root.destroy()

root.protocol("WM_DELETE_WINDOW", on_close)
root.mainloop()
